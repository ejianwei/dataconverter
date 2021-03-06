# coding=utf-8
import os
import sys
import re
from openpyxl import Workbook
from openpyxl import load_workbook

import ding_conf

if sys.getdefaultencoding() != 'utf-8':
    reload(sys)
    sys.setdefaultencoding('utf-8')


def convert_course_read(file, classNo, term):

    print ("Converting course read, file: " + file + " classNo: " + classNo + " term: " + str(term))

    workbook = load_workbook(filename=file, read_only=True ,data_only=True )
    sheet = workbook.active
    
    #保存表头
    header = ()
    #只遍历一次，把表头（excel第一行）保存到header
    for row in sheet.iter_rows(min_row=1,
                                 max_row=1,
                                 values_only=True):
        header = row

    #处理数据
    output_data_list = []
    #输出表的表头
    output_data_list.append(["姓名","学号","班级编号","学期","课程","文档完成","视频完成","课后题完成"])
    
    for row in sheet.iter_rows(min_row=2, values_only=True):


        if not row[0]:
            continue

        #如果学员未填写
        if not row[5] or row[5].find("是") == -1:
            print ("Skip processing for: " +  row[0].decode('utf-8'))
            continue

        noIdx = len(row) - 1 #倒数第一列，学号
        nameIdx = noIdx - 1 # 倒数第二列，姓名

        #导出的表格，第6列开始是学修数据
    	for idx in range(6, nameIdx):
            course_read = []
            course_read.append(row[nameIdx]) #姓名
            course_read.append(row[noIdx]) #学号
            course_read.append(classNo) #班级编号
            course_read.append(term) #学期
            course_read.append(header[idx]) #课程
            if not row[idx]:
                print "ERROR: " + row[0] + "column: " + str(idx) 
                return

    	    if row[idx].find("文档完成") != -1:
                course_read.append("是")
            else:
                course_read.append("否")

            if row[idx].find("视频完成") != -1:
                course_read.append("是")
            else:
                course_read.append("否")

            if row[idx].find("课后题完成") != -1:
                course_read.append("是")
            else:
                course_read.append("否")

            output_data_list.append(course_read)

    #保存数据到新文件
    filename = ding_conf.output_path + "\学修数据" + classNo + ".xlsx"

    new_workbook = Workbook()
    new_sheet = new_workbook.active


    for output_data in output_data_list:
        new_sheet.append(output_data)

    new_workbook.save(filename=filename.decode('utf-8'))
    print ("write output file: " + filename.decode('utf-8'))

def convert_attendance(file, classNo, term):
    print ("Converting att, file: " + file + " classNo: " + classNo + " term: " + str(term))

    workbook = load_workbook(filename=file, read_only=True ,data_only=True )
    sheet = workbook.active
    
    #保存表头
    header = ()
    #只遍历一次，把表头（excel第一行）保存到header
    for row in sheet.iter_rows(min_row=1,
                                 max_row=1,
                                 values_only=True):
        header = row

    # for idx in range(0, len(header)):
    #     print header[idx]
    # return
    #处理数据
    output_data_list = []
    #输出表的表头
    output_data_list.append(["姓名","学号","班级编号","学期","周数","出勤"])
    

    for row in sheet.iter_rows(min_row=3,
                                 values_only=True):

        #跳过空行。没有姓名的行，视为空行
        if (not row[0]):
            continue

        #周数3+，名字0，学号1
        for idx in range(3, len(row)):
            if not header[idx]:
                continue
            attendance = []
            attendance.append(row[0]) #姓名
            attendance.append(row[1]) #学号
            attendance.append(classNo) #班级编号
            attendance.append(term) #学期
            attendance.append(re.search("\d+", str(header[idx])).group(0))#提取表头第几周中的数字
            #出勤
            if row[idx] != "出勤" and row[idx] != "请假" and row[idx] != "迟到"and row[idx] != "旷课":
                attendance.append("无") 
            else:
                attendance.append(row[idx])

            output_data_list.append(attendance)


    #保存数据到新文件
    filename = ding_conf.output_path + "\出勤数据" + classNo + ".xlsx"

    new_workbook = Workbook()
    new_sheet = new_workbook.active


    for output_data in output_data_list:
        new_sheet.append(output_data)

    new_workbook.save(filename=filename.decode('utf-8'))
    print ("write output file: " + filename.decode('utf-8'))

def main(argv):

    term = ding_conf.term #第一学期

    if argv[0] == "att":
        for root, dirs, files in os.walk(ding_conf.input_att_dir):
            # 遍历文件
            for f in files:
                classNo = re.search("[a-zA-Z0-9-]+", f).group(0);
                convert_attendance(os.path.join(root, f), classNo, term)
                
    elif argv[0] == "course":
        for root, dirs, files in os.walk(ding_conf.input_course_dir):
            # 遍历文件
            for f in files:
                classNo = re.search("[a-zA-Z0-9-]+", f).group(0);
                convert_course_read(os.path.join(root, f), classNo, term)
    else:
        print ("Error: Wrong Parameter: " + argv[0])

if __name__ == "__main__":
	main(sys.argv[1:])
   
