# coding=utf-8
import os
import sys
import re
from openpyxl import Workbook
from openpyxl import load_workbook
import ding_conf





if sys.getdefaultencoding() != 'utf-8':
    reload(sys)
    sys.setdefaultencoding('utf-8')




def scan_header(file, standard_file):
    
    workbook = load_workbook(filename=file, read_only=True ,data_only=True )
    sheet = workbook.active
    
    sWorkbook = load_workbook(filename=standard_file, read_only=True ,data_only=True )
    sSheet = sWorkbook.active

    #保存表头
    header = ()
    #只遍历一次，把表头（excel第一行）保存到header
    for row in sheet.iter_rows(min_row=1,
                                 max_row=1,
                                 values_only=True):
        header = row

    #保存表头
    sheader = ()
    #只遍历一次，把表头（excel第一行）保存到header
    for row in sSheet.iter_rows(min_row=1,
                                 max_row=1,
                                 values_only=True):
        sheader = row

    for idx in range(0, len(sheader)):
        if sheader[idx] != header[idx]:
            print ("Error: invalid header : " + str(header[idx]).decode('utf-8'))
            print ("    ==> correct header : " + str(sheader[idx]).decode('utf-8'))



def scan_course_data(file):
    workbook = load_workbook(filename=file, read_only=True ,data_only=True )
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):

        if not row[0]:
            continue

        #如果学员未填写
        if row[5].find("是") == -1:
            #print ("Skip processing for: " +  row[0].decode('utf-8'))
            continue

        noIdx = len(row) - 1 #倒数第一列，学号

        # if not row[noIdx]:
        #     print (row[noIdx-1] + " 学号错误: [" +  str(row[noIdx]) +"]")

        if not re.match(ding_conf.studentNO_format, str(row[noIdx]), flags=0):
            print (row[noIdx-1] + " 学号错误: [" +  str(row[noIdx]) +"]")


        nameIdx = noIdx - 1 # 倒数第二列，姓名

        #导出的表格，第6列开始是学修数据
        for idx in range(6, nameIdx):

            if row[idx] == "-":
                continue

            if not row[idx]: 
                print (row[noIdx-1] + "答案错误：[none], coloumn: " + str(idx))
                continue

            answers = row[idx].split(",")
            for answer in answers:
                if answer != "文档完成" and answer != "视频完成" and answer != "课后题完成":
                    print (row[noIdx-1] + "答案错误：[" + row[idx] + "]")



def scan_att_data(file):
    workbook = load_workbook(filename=file, read_only=True ,data_only=True )
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=3, values_only=True):

        #跳过空行。没有姓名的行，视为空行
        if (not row[0]):
            continue
        
        # if not row[1]:
        #     print (row[0] + " 学号错误: [" +  str(row[1]) +"]")

        if not re.match(ding_conf.studentNO_format, str(row[1]), flags=0):
            print (row[0] + " 学号错误: [" +  str(row[1]) +"]")


        for idx in range(3, 26):
            
            if not row[idx]:
                continue
            attdata = str(row[idx])
            if attdata == "　" or attdata == "出勤" or attdata == "请假" or attdata == "迟到" or attdata == "旷课":
                continue
            print (row[0] + " 出勤数据错误，第" + str(idx) + "列: "+ "[" + attdata  +"]")
        

   
def main(argv):

    if argv[0] == "att":
        for root, dirs, files in os.walk(ding_conf.input_att_dir):
            # 遍历文件
            for f in files:
                print ("Scanning file: " + f)

                scan_header(os.path.join(root, f), ding_conf.att_standard_path)
                scan_att_data(os.path.join(root, f))
                
    elif argv[0] == "course":
        for root, dirs, files in os.walk(ding_conf.input_course_dir):
            # 遍历文件
            for f in files:
                print ("Scanning file: " + f)
                #scan_header(os.path.join(root, f), ".\standard\course_standard.xlsx")
                scan_header(os.path.join(root, f), ding_conf.course_standard_path)
                
                scan_course_data(os.path.join(root, f))
    else:
        print ("Error: Wrong Parameter: " + argv[0])

if __name__ == "__main__":
    main(sys.argv[1:])